#!/usr/bin/env python3
"""
Markdown 测试用例转换为 XLSX 格式脚本（本地版）

功能：
1. 解析 Markdown 格式的测试用例文件
2. 提取测试用例表格数据
3. 生成本地 XLSX 格式（无需 MeterSphere 配置）
4. 支持批量转换

与 md_to_xlsx.py 的区别：
- 不依赖 MS_USERNAME、MS_PASSWORD、MS_PROJECT、MS_DIR、MS_VERSION 等环境变量
- 所属模块、标签、责任人使用默认值，可直接本地查看和编辑

使用方法：
python scripts/md_to_xlsx_local.py <markdown文件路径> [输出文件路径]

示例：
python scripts/md_to_xlsx_local.py output/atm/出金命中规则_testcase.md
python scripts/md_to_xlsx_local.py output/atm/出金命中规则_testcase.md output/atm/出金命中规则.xlsx
"""
import sys
import re
from pathlib import Path

try:
    import yaml
except ImportError:
    yaml = None
from typing import List, Dict, Optional

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment
from openpyxl.utils.dataframe import dataframe_to_rows


class TestCaseParser:
    """测试用例解析器"""

    def __init__(self, md_file: Path):
        self.md_file = md_file
        self.metadata = {}
        self.testcases = []

    def parse(self) -> Dict:
        """解析 Markdown 文件"""
        with open(self.md_file, 'r', encoding='utf-8') as f:
            content = f.read()

        # 提取元数据
        self.metadata = self._extract_metadata(content)

        # 提取测试用例
        self.testcases = self._extract_testcases(content)

        return {
            'metadata': self.metadata,
            'testcases': self.testcases
        }

    def _extract_metadata(self, content: str) -> Dict:
        """提取 YAML 元数据"""
        yaml_pattern = r'^---\s*\n(.*?)\n---'
        match = re.search(yaml_pattern, content, re.DOTALL | re.MULTILINE)

        if match and yaml is not None:
            try:
                return yaml.safe_load(match.group(1))
            except yaml.YAMLError:
                pass

        return {}

    def _extract_testcases(self, content: str) -> List[Dict]:
        """提取测试用例表格"""
        testcases = []

        table_pattern = r'\|(.+?)\|\s*\n\|([\s\-:]+\|)+\s*\n((?:\|.+\|\s*\n?)+)'

        for match in re.finditer(table_pattern, content, re.MULTILINE):
            header_line = match.group(1)
            data_lines = match.group(3)

            headers = [h.strip() for h in header_line.split('|') if h.strip()]

            is_new_format = '用例名称' in headers and '所属模块' in headers
            is_old_format = '用例ID' in headers or '用例标题' in headers

            if not (is_new_format or is_old_format):
                continue

            for line in data_lines.strip().split('\n'):
                if line.strip():
                    cells = [c.strip() for c in line.split('|')[1:-1]]

                    if len(cells) == len(headers):
                        testcase = dict(zip(headers, cells))
                        testcase = self._clean_testcase(testcase)
                        testcases.append(testcase)

        return testcases

    def _clean_testcase(self, testcase: Dict) -> Dict:
        """清理测试用例数据"""
        cleaned = {}
        for key, value in testcase.items():
            value = re.sub(r'<br\s*/?>', '\n', value)
            value = re.sub(r'<[^>]+>', '', value)
            value = value.strip()
            cleaned[key] = value
        return cleaned


class XLSXConverter:
    """XLSX 转换器（本地版，无需 MS 配置）"""

    def __init__(self, testcases: List[Dict], metadata: Dict):
        self.testcases = testcases
        self.metadata = metadata

    def _convert_old_to_new_format(self, df: 'pd.DataFrame') -> 'pd.DataFrame':
        """将旧格式转换为新格式"""
        column_mapping = {
            '用例标题': '用例名称',
            '测试步骤': '步骤描述',
            '优先级': '用例等级',
            '用例类型': '备注',
        }

        for old_col, new_col in column_mapping.items():
            if old_col in df.columns and new_col not in df.columns:
                df.rename(columns={old_col: new_col}, inplace=True)

        if '用例ID' in df.columns and '备注' in df.columns:
            df['备注'] = df.apply(
                lambda row: f"用例ID: {row['用例ID']}; {row.get('备注', '')}" if row.get('用例ID') else row.get('备注', ''),
                axis=1
            )

        if '关联需求' in df.columns:
            df['备注'] = df.apply(
                lambda row: f"{row.get('备注', '')}; 关联需求: {row['关联需求']}" if row.get('关联需求') else row.get('备注', ''),
                axis=1
            )

        if '所属模块' not in df.columns and self.metadata.get('module'):
            module_name = self.metadata.get('module', '')
            module_names = {
                'sbl': '港股融券SBL',
                'clearing': '清算模块',
                'billing': '计费模块',
            }
            module_display = module_names.get(module_name, module_name.upper())
            df['所属模块'] = f"/{module_display}/功能测试"

        return df

    @staticmethod
    def split_and_expand(row):
        """分割步骤描述和预期结果并展开为多行"""
        pre_conditions = row["步骤描述"].split("；")
        step_descriptions = row["预期结果"].split("；")

        max_length = max(len(pre_conditions), len(step_descriptions))
        pre_conditions.extend([""] * (max_length - len(pre_conditions)))
        step_descriptions.extend([""] * (max_length - len(step_descriptions)))

        rows = []
        for i in range(max_length):
            if i == 0:
                raw_row = row.to_dict()
                raw_row.update(**{
                    "步骤描述": pre_conditions[i],
                    "预期结果": step_descriptions[i]
                })
                rows.append(raw_row)
            else:
                rows.append({
                    "步骤描述": pre_conditions[i],
                    "预期结果": step_descriptions[i]
                })
        return rows

    @staticmethod
    def merge_cells(ws, col_index):
        """合并指定列的单元格"""
        start_row = None
        prev_value = None

        for row in range(2, ws.max_row + 1):
            current_value = ws.cell(row=row, column=col_index).value

            if not pd.isnull(current_value):
                if start_row is not None and prev_value is not None:
                    ws.merge_cells(start_row=start_row, end_row=row - 1, start_column=col_index, end_column=col_index)
                    ws.cell(row=start_row, column=col_index).alignment = Alignment(vertical="center",
                                                                                   horizontal="center")
                start_row = row
            elif start_row is None:
                start_row = row

            prev_value = current_value

        if start_row is not None and prev_value is not None:
            ws.merge_cells(start_row=start_row, end_row=ws.max_row, start_column=col_index, end_column=col_index)
            ws.cell(row=start_row, column=col_index).alignment = Alignment(vertical="center", horizontal="center")

    def convert(self, output_file: Path):
        """转换为 XLSX 格式（使用本地默认值，无需 MS 配置）"""
        if not self.testcases:
            print("⚠️  警告: 没有找到测试用例数据")
            return

        # 创建 DataFrame
        df = pd.DataFrame(self.testcases)
        df['是否采纳'] = ''
        # 保留原始所属模块，不添加前缀
        if '所属模块' not in df.columns:
            df['所属模块'] = "/功能测试"
        if '标签' in df.columns:
            df['标签'] = df['标签'].apply(lambda x: f"AI,{x}" if x else "AI,本地")
        else:
            df['标签'] = 'AI'
        # 保留原始责任人，不覆盖
        if '责任人' not in df.columns:
            df['责任人'] = ''
        df['编辑模式'] = 'STEP'
        df['用例状态'] = '进行中'

        required_columns = [
            '是否采纳',
            '用例名称',
            '所属模块',
            '标签',
            '前置条件',
            '步骤描述',
            '预期结果',
            '编辑模式',
            '备注',
            '用例状态',
            '责任人',
            '用例等级',
        ]

        if '用例ID' in df.columns or '用例标题' in df.columns:
            print("   检测到旧格式，尝试转换为新格式...")
            df = self._convert_old_to_new_format(df)

        for col in required_columns:
            if col not in df.columns:
                df[col] = 'AI' if col == '标签' else ''

        existing_columns = [col for col in required_columns if col in df.columns]
        df = df[existing_columns]

        output_file.parent.mkdir(parents=True, exist_ok=True)

        df['前置条件'] = df['前置条件'].apply(func=lambda x: "\n".join(x.split('；')) if isinstance(x, str) else '')

        expanded_rows = []
        case_ranges = []
        current_row = 0

        for _, row in df.iterrows():
            case_rows = self.split_and_expand(row)
            start_row = current_row
            expanded_rows.extend(case_rows)
            current_row += len(case_rows)
            end_row = current_row - 1
            case_ranges.append((start_row, end_row))

        new_df = pd.DataFrame(expanded_rows)

        wb = Workbook()
        ws = wb.active
        ws.title = '测试用例'

        for row in dataframe_to_rows(new_df, index=False, header=True):
            ws.append(row)

        step_col_index = None
        result_col_index = None
        for col_index, col_name in enumerate(new_df.columns, 1):
            if col_name == '步骤描述':
                step_col_index = col_index
            elif col_name == '预期结果':
                result_col_index = col_index
            else:
                self.merge_cells(ws, col_index)

        from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

        column_widths = {
            '是否采纳': 12,
            '用例名称': 35,
            '所属模块': 30,
            '用例等级': 12,
            '备注': 30,
            '前置条件': 30,
            '步骤描述': 45,
            '预期结果': 40,
            '标签': 10,
            '编辑模式': 12,
            '用例状态': 12,
            '责任人': 12,
        }

        for col_index, col_name in enumerate(new_df.columns, 1):
            col_letter = ws.cell(row=1, column=col_index).column_letter
            width = column_widths.get(col_name, 15)
            ws.column_dimensions[col_letter].width = width

        header_fill = PatternFill(start_color="4472C4", end_color="4472C4", fill_type="solid")
        header_font = Font(color="FFFFFF", bold=True, size=11)
        header_alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        thin_border = Border(
            left=Side(style='thin', color='000000'),
            right=Side(style='thin', color='000000'),
            top=Side(style='thin', color='000000'),
            bottom=Side(style='thin', color='000000')
        )

        for cell in ws[1]:
            cell.fill = header_fill
            cell.font = header_font
            cell.alignment = header_alignment
            cell.border = thin_border

        zebra_color_1 = PatternFill(start_color="F2F2F2", end_color="F2F2F2", fill_type="solid")
        zebra_color_2 = PatternFill(start_color="E8F4F8", end_color="E8F4F8", fill_type="solid")

        for row_idx, row in enumerate(ws.iter_rows(min_row=2, max_row=ws.max_row), start=2):
            ws.row_dimensions[row_idx].height = None

            data_row_idx = row_idx - 2
            case_index = None
            for idx, (start, end) in enumerate(case_ranges):
                if start <= data_row_idx <= end:
                    case_index = idx
                    break

            bg_fill = None if case_index is not None and case_index % 2 == 0 else zebra_color_2

            for col_idx, cell in enumerate(row, start=1):
                cell.border = thin_border

                if bg_fill:
                    cell.fill = bg_fill

                if col_idx == step_col_index or col_idx == result_col_index:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)
                elif new_df.columns[col_idx - 1] in ['标签', '编辑模式', '用例状态', '责任人', '用例等级']:
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                else:
                    cell.alignment = Alignment(horizontal='left', vertical='top', wrap_text=True)

        ws.freeze_panes = 'A2'

        wb.save(output_file)

        print(f"✅ 转换成功！")
        print(f"   输出文件: {output_file}")
        print(f"   测试用例数: {len(self.testcases)}")


def convert_md_to_xlsx(md_file: Path, output_file: Optional[Path] = None):
    """转换 Markdown 到 XLSX（本地版）"""
    if output_file is None:
        output_file = md_file.parent / md_file.name.replace('.md', '.xlsx')

    print(f"📄 解析 Markdown 文件: {md_file.name}")

    parser = TestCaseParser(md_file)
    data = parser.parse()

    print(f"   找到 {len(data['testcases'])} 条测试用例")

    if not data['testcases']:
        print("⚠️  警告: 未找到测试用例表格")
        print("   请确保 Markdown 文件包含标准的测试用例表格")
        return

    print(f"\n📊 转换为 XLSX 格式（本地版）...")
    converter = XLSXConverter(data['testcases'], data['metadata'])
    converter.convert(output_file)


def main():
    """主函数"""
    if len(sys.argv) < 2:
        print("使用方法: python scripts/md_to_xlsx_local.py <markdown文件路径> [输出文件路径]")
        print("\n示例:")
        print("  python scripts/md_to_xlsx_local.py output/atm/出金命中规则_testcase.md")
        print("  python scripts/md_to_xlsx_local.py output/atm/出金命中规则_testcase.md output/atm/出金命中规则.xlsx")
        print("\n说明: 本脚本为本地版，无需配置 MeterSphere 环境变量")
        sys.exit(1)

    md_file = Path(sys.argv[1])

    if not md_file.exists():
        print(f"❌ 错误: 文件不存在: {md_file}")
        sys.exit(1)

    if md_file.suffix != '.md':
        print(f"❌ 错误: 不是 Markdown 文件: {md_file}")
        sys.exit(1)

    output_file = Path(sys.argv[2]) if len(sys.argv) > 2 else None

    try:
        convert_md_to_xlsx(md_file, output_file)
        print(f"\n🎉 转换完成！")
    except Exception as e:
        print(f"\n❌ 错误: {e}")
        import traceback
        traceback.print_exc()
        sys.exit(1)


if __name__ == "__main__":
    main()
